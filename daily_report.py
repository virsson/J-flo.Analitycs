"""
Единый ежедневный отчёт J-flo.ru: Google Search Console + Яндекс.Вебмастер.

Что делает:
  1. Выгружает данные GSC за T-3 (с динамикой к T-4).
  2. Выгружает данные Яндекса за 7 дней (T-9 .. T-3).
  3. Обновляет накапливающий keywords.xlsx (Google).
  4. Пишет сводный XLSX с 5 листами:
       Главный | Google | Yandex | Ошибки | Рекомендации
  5. Шлёт в Telegram краткую сводку: GOOGLE + YANDEX (только хосты с трафиком).

Использование:
    python daily_report.py                 # за T-3
    python daily_report.py 2026-04-15      # за конкретную дату (Google)
"""
import os
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
sys.path.insert(0, os.path.join(BASE_DIR, 'yandex_webmaster'))

from search_console import get_service as get_gsc_service
from search_console import google_report as gsc
from yandex_webmaster import yandex_data as yw
from telegram_bot import notifier


REPORTS_DIR = os.path.join(BASE_DIR, 'reports')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='D9E1F2')


def resolve_gsc_date(argv):
    if len(argv) > 1:
        return argv[1]
    return (date.today() - timedelta(days=3)).isoformat()


def _apply_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def _autosize(ws, max_width=60):
    for col_idx, col in enumerate(ws.columns, 1):
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, max_width)


def write_main_sheet(wb, gsc_grouped, yw_data, target_date):
    ws = wb.create_sheet('Главный')
    note = (
        f'G = Google Search Console за {target_date} (1 день)    '
        f'Y = Яндекс Вебмастер за {yw_data["date_from"]}..{yw_data["date_to"]} (7 дней)'
    )
    ws.append([note])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws['A1'].font = Font(italic=True, color='555555')
    ws['A1'].alignment = Alignment(horizontal='left')
    ws.append([
        'host',
        'G_запросы', 'G_клики', 'G_показы', 'G_позиция',
        'Y_запросы', 'Y_клики', 'Y_показы', 'Y_позиция', 'Y_CTR',
    ])
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    yw_by_host = {h['host']: h for h in yw_data['hosts']}
    all_hosts = set(gsc_grouped.keys()) | set(yw_by_host.keys())
    sorted_hosts = sorted(all_hosts, key=yw.host_sort_key)

    totals = {k: 0 for k in ['gq', 'gc', 'gi', 'yq', 'yc', 'ys']}
    g_pos_weighted = 0
    y_pos_weighted = 0

    for host in sorted_hosts:
        g_rows = gsc_grouped.get(host, [])
        g_agg = gsc.aggregate(g_rows) if g_rows else {
            'queries': 0, 'clicks': 0, 'impressions': 0, 'position': 0,
        }
        y_host = yw_by_host.get(host)
        y_agg = yw.aggregate(y_host['queries']) if y_host else {
            'queries': 0, 'clicks': 0, 'shows': 0, 'ctr': 0, 'position': 0,
        }

        ws.append([
            host,
            g_agg['queries'], g_agg['clicks'], g_agg['impressions'],
            round(g_agg['position'], 2) if g_agg['queries'] else '',
            y_agg['queries'], y_agg['clicks'], y_agg['shows'],
            round(y_agg['position'], 2) if y_agg['queries'] else '',
            round(y_agg['ctr'], 4) if y_agg['shows'] else '',
        ])
        totals['gq'] += g_agg['queries']
        totals['gc'] += g_agg['clicks']
        totals['gi'] += g_agg['impressions']
        totals['yq'] += y_agg['queries']
        totals['yc'] += y_agg['clicks']
        totals['ys'] += y_agg['shows']
        g_pos_weighted += g_agg['position'] * g_agg['impressions']
        y_pos_weighted += y_agg['position'] * y_agg['shows']

    g_avg = g_pos_weighted / totals['gi'] if totals['gi'] else 0
    y_avg = y_pos_weighted / totals['ys'] if totals['ys'] else 0
    y_ctr = totals['yc'] / totals['ys'] if totals['ys'] else 0

    ws.append([
        'ИТОГО',
        totals['gq'], totals['gc'], totals['gi'], round(g_avg, 2),
        totals['yq'], totals['yc'], totals['ys'], round(y_avg, 2),
        round(y_ctr, 4),
    ])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    _autosize(ws)


def write_google_sheet(wb, gsc_grouped):
    ws = wb.create_sheet('Google')
    _apply_headers(ws, [
        'host', 'query', 'page', 'clicks', 'impressions', 'ctr', 'position',
    ])
    for host in sorted(gsc_grouped.keys(), key=yw.host_sort_key):
        rows = sorted(gsc_grouped[host], key=lambda r: r.get('clicks', 0), reverse=True)
        for r in rows:
            query, page = r['keys']
            ws.append([
                host, query, page,
                r.get('clicks', 0), r.get('impressions', 0),
                round(r.get('ctr', 0), 4), round(r.get('position', 0), 2),
            ])
    _autosize(ws, max_width=80)


def write_yandex_sheet(wb, yw_data):
    ws = wb.create_sheet('Yandex')
    _apply_headers(ws, [
        'host', 'query', 'clicks', 'shows', 'ctr',
        'show_position', 'click_position',
    ])
    hosts_sorted = sorted(yw_data['hosts'], key=lambda h: yw.host_sort_key(h['host']))
    for h in hosts_sorted:
        qs = sorted(h['queries'], key=lambda q: q['clicks'], reverse=True)
        for q in qs:
            ws.append([
                h['host'], q['query'],
                q['clicks'], q['shows'], round(q['ctr'], 4),
                round(q['show_position'], 2) if q['show_position'] else '',
                round(q['click_position'], 2) if q['click_position'] else '',
            ])
    _autosize(ws, max_width=80)


def write_errors_sheet(wb, sitemaps, sitemap_problems, yw_data):
    ws = wb.create_sheet('Ошибки')
    _apply_headers(ws, ['source', 'host', 'type', 'details'])

    # Google sitemap — только реальные ошибки/предупреждения
    for s in sitemaps:
        path = s.get('path', '')
        errors = int(s.get('errors', 0) or 0)
        warnings = int(s.get('warnings', 0) or 0)
        if errors > 0 or warnings > 0:
            ws.append([
                'Google Sitemap', path,
                f'errors={errors}, warnings={warnings}',
                f'last_downloaded={s.get("lastDownloaded", "") or "не скачан"}',
            ])

    # Yandex host problems — только FATAL/CRITICAL (RECOMMENDATION/POSSIBLE_PROBLEM = шум)
    SEVERE = {'FATAL', 'CRITICAL', 'ERROR'}
    for h in yw_data['hosts']:
        problems = h['summary'].get('site_problems') or {}
        if isinstance(problems, dict):
            for key, val in problems.items():
                try:
                    count = int(val)
                except (TypeError, ValueError):
                    count = 0
                if key in SEVERE and count > 0:
                    ws.append(['Yandex', h['host'], key, f'count={count}'])
        if 'error' in h['summary']:
            ws.append(['Yandex', h['host'], 'api_error', h['summary']['error']])
        if 'queries_error' in h['summary']:
            ws.append(['Yandex', h['host'], 'queries_api_error', h['summary']['queries_error']])

    if ws.max_row == 1:
        ws.append(['(нет ошибок)', '', '', ''])
    _autosize(ws, max_width=80)


def write_recommendations_sheet(wb, recs):
    ws = wb.create_sheet('Рекомендации')
    _apply_headers(ws, [
        'source', 'type', 'host', 'query', 'page',
        'clicks', 'impressions', 'ctr', 'position', 'reason',
    ])
    for r in recs:
        host = gsc.get_host(r['page'])
        ws.append([
            'Google', r['type'], host, r['query'], r['page'],
            r['clicks'], r['impressions'],
            round(r['ctr'], 4), round(r['position'], 2), r['reason'],
        ])
    if ws.max_row == 1:
        ws.append(['(рекомендаций нет)', '', '', '', '', '', '', '', '', ''])
    _autosize(ws, max_width=80)


def write_xlsx(target_date, gsc_grouped, gsc_sitemaps, gsc_sitemap_problems,
               gsc_recs, yw_data):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f'daily_report_{target_date}.xlsx')
    wb = Workbook()
    wb.remove(wb.active)

    write_main_sheet(wb, gsc_grouped, yw_data, target_date)
    write_google_sheet(wb, gsc_grouped)
    write_yandex_sheet(wb, yw_data)
    write_errors_sheet(wb, gsc_sitemaps, gsc_sitemap_problems, yw_data)
    write_recommendations_sheet(wb, gsc_recs)

    wb.save(path)
    return path


def main():
    target_date = resolve_gsc_date(sys.argv)

    # === Google ===
    print(f'=== GOOGLE Search Console за {target_date} ===')
    service = get_gsc_service()
    rows = gsc.fetch_rows(service, target_date)
    print(f'Получено строк: {len(rows)}')
    if not rows:
        print('Нет данных GSC за эту дату — прерываю.')
        return

    prev = gsc.prev_date(target_date)
    prev_rows = gsc.fetch_rows(service, prev)
    print(f'Предыдущий день ({prev}): {len(prev_rows)} строк')

    grouped = gsc.group_by_host(rows)
    prev_grouped = gsc.group_by_host(prev_rows)
    print(f'GSC субдоменов: {len(grouped)}')

    gsc.update_keywords_xlsx(rows, target_date)

    sitemaps = gsc.fetch_sitemap_status(service)
    sitemap_problems = gsc.find_sitemap_issues(sitemaps)
    print(f'Sitemap-проблем: {len(sitemap_problems)}/{len(sitemaps)}')

    recs = gsc.build_recommendations(rows, prev_rows)
    print(f'Рекомендаций: {len(recs)}')

    # === Yandex (за 7 дней) ===
    print()
    yw_data = yw.fetch_data()

    # === XLSX ===
    print()
    xlsx_path = write_xlsx(
        target_date, grouped, sitemaps, sitemap_problems, recs, yw_data,
    )
    print(f'XLSX: {xlsx_path}')

    # === Telegram ===
    critical = gsc.build_critical_block(sitemap_problems, grouped, prev_grouped)
    google_section = gsc.build_summary(target_date, grouped, prev_grouped, critical)
    yandex_section = yw.build_summary_text(yw_data)

    try:
        notifier.send_report([google_section, yandex_section])
        print('Telegram: сводка отправлена')
    except Exception as e:
        print(f'Telegram: ошибка — {e}')


if __name__ == '__main__':
    main()

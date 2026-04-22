"""
Ежедневный отчёт по j-flo.ru из Google Search Console — с разбивкой по
субдоменам.

Что делает:
  1. Выгружает статистику за указанную дату (по умолчанию — 3 дня назад,
     т.к. Search Console обновляет данные с задержкой 2-3 дня).
  2. Группирует строки по субдомену (hostname страницы входа).
  3. Для каждого субдомена сохраняет отдельный CSV:
         reports/daily_report_YYYY-MM-DD_<host>.csv
  4. Обновляет накапливающий keywords.xlsx (уникальные пары ключ+страница).
  5. Шлёт в Telegram краткую сводку: секции идут в фиксированном порядке —
     сначала j-flo.ru, затем sankt-peterburg.j-flo.ru, дальше остальные по
     алфавиту.

Использование:
    python daily_report.py                 # за 3 дня назад
    python daily_report.py 2026-04-15      # за конкретную дату
"""
import csv
import html
import os
import sys
from datetime import date, timedelta
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

from search_console import get_service, list_sitemaps

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from telegram_bot import notifier


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(BASE_DIR, 'reports')
KEYWORDS_FILE = os.path.join(BASE_DIR, 'keywords.xlsx')

SITE_URL = 'sc-domain:j-flo.ru'
CSV_HEADERS = ['query', 'page', 'clicks', 'impressions', 'ctr', 'position']
XLSX_HEADERS = [
    'query', 'page', 'first_seen', 'last_seen',
    'total_clicks', 'total_impressions', 'last_position',
]

PRIORITY_HOSTS = ['j-flo.ru', 'sankt-peterburg.j-flo.ru']

# Пороги для рекомендаций / критических предупреждений
LOW_CTR_MIN_IMPRESSIONS = 50
LOW_CTR_THRESHOLD = 0.02
PAGE_2_MIN = 11
PAGE_2_MAX = 20
POSITION_DROP_DELTA = 3.0
CRITICAL_CLICKS_DROP_RATIO = 0.5
CRITICAL_CLICKS_MIN_BASELINE = 10


def resolve_date(argv):
    if len(argv) > 1:
        return argv[1]
    return (date.today() - timedelta(days=3)).isoformat()


def get_host(url):
    host = (urlparse(url).hostname or '').lower()
    if host.startswith('www.'):
        host = host[4:]
    return host


def host_sort_key(host):
    if host in PRIORITY_HOSTS:
        return (0, PRIORITY_HOSTS.index(host))
    return (1, host)


def fetch_rows(service, target_date):
    """Запрашиваем все строки за день. Постранично через startRow."""
    rows = []
    start_row = 0
    page_size = 25000
    while True:
        body = service.searchanalytics().query(siteUrl=SITE_URL, body={
            'startDate': target_date,
            'endDate': target_date,
            'dimensions': ['query', 'page'],
            'rowLimit': page_size,
            'startRow': start_row,
        }).execute()
        batch = body.get('rows', [])
        rows.extend(batch)
        if len(batch) < page_size:
            break
        start_row += page_size
    return rows


def prev_date(target_date):
    return (date.fromisoformat(target_date) - timedelta(days=1)).isoformat()


def aggregate(rows):
    n = len(rows)
    clicks = sum(r.get('clicks', 0) for r in rows)
    imps = sum(r.get('impressions', 0) for r in rows)
    pos = sum(r.get('position', 0) for r in rows) / n if n else 0
    return {'queries': n, 'clicks': clicks, 'impressions': imps, 'position': pos}


def fmt_int_delta(curr, prev):
    d = curr - prev
    if d == 0:
        return '(=)'
    return f'(+{d})' if d > 0 else f'({d})'


def fmt_pos_delta(curr, prev):
    d = curr - prev
    if abs(d) < 0.05:
        return '(=)'
    return f'(+{d:.2f})' if d > 0 else f'({d:.2f})'


def group_by_host(rows):
    """{host: [row, ...]} в нужном порядке (приоритетные сверху)."""
    groups = {}
    for row in rows:
        page = row['keys'][1]
        host = get_host(page) or '(unknown)'
        groups.setdefault(host, []).append(row)
    return dict(sorted(groups.items(), key=lambda kv: host_sort_key(kv[0])))


def write_csv(rows, target_date, host):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f'daily_report_{target_date}_{host}.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(CSV_HEADERS)
        for row in rows:
            query, page = row['keys']
            writer.writerow([
                query, page,
                row.get('clicks', 0),
                row.get('impressions', 0),
                f"{row.get('ctr', 0):.4f}",
                f"{row.get('position', 0):.2f}",
            ])
    return path


def update_keywords_xlsx(rows, target_date):
    """Накапливает уникальные (query, page) с агрегированной статистикой."""
    if os.path.exists(KEYWORDS_FILE):
        wb = load_workbook(KEYWORDS_FILE)
        ws = wb.active
        existing = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            existing[(r[0], r[1])] = list(r)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'keywords'
        ws.append(XLSX_HEADERS)
        existing = {}

    for row in rows:
        query, page = row['keys']
        clicks = row.get('clicks', 0)
        impressions = row.get('impressions', 0)
        position = row.get('position', 0)
        key = (query, page)
        if key in existing:
            rec = existing[key]
            rec[3] = target_date
            rec[4] = (rec[4] or 0) + clicks
            rec[5] = (rec[5] or 0) + impressions
            rec[6] = round(position, 2)
        else:
            existing[key] = [
                query, page, target_date, target_date,
                clicks, impressions, round(position, 2),
            ]

    # сортировка: по приоритету хоста, потом по clicks убыв.
    sorted_recs = sorted(
        existing.values(),
        key=lambda r: (host_sort_key(get_host(r[1])), -(r[4] or 0)),
    )

    ws.delete_rows(2, ws.max_row)
    for rec in sorted_recs:
        ws.append(rec)

    wb.save(KEYWORDS_FILE)
    return KEYWORDS_FILE, len(existing)


def fetch_sitemap_status(service):
    """Статус всех зарегистрированных sitemap-ов."""
    return list_sitemaps(service, SITE_URL)


def find_sitemap_issues(sitemaps):
    """Sitemap-ы с реальными ошибками/предупреждениями (не считаем «не скачан»)."""
    problems = []
    for s in sitemaps:
        errors = int(s.get('errors', 0) or 0)
        warnings = int(s.get('warnings', 0) or 0)
        if errors > 0 or warnings > 0:
            problems.append({
                'path': s.get('path', ''),
                'errors': errors,
                'warnings': warnings,
                'last_submitted': s.get('lastSubmitted', ''),
                'last_downloaded': s.get('lastDownloaded', '') or '',
                'is_pending': s.get('isPending', False),
            })
    return problems


def write_errors_csv(sitemaps, sitemap_problems, target_date):
    """Все sitemap-ы + пометка is_problem. Пустой CSV — тоже нормально."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f'errors_{target_date}.csv')
    problem_paths = {p['path'] for p in sitemap_problems}
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'sitemap', 'errors', 'warnings', 'last_submitted',
            'last_downloaded', 'is_pending', 'is_problem',
        ])
        for s in sitemaps:
            w.writerow([
                s.get('path', ''),
                s.get('errors', 0),
                s.get('warnings', 0),
                s.get('lastSubmitted', ''),
                s.get('lastDownloaded', ''),
                s.get('isPending', False),
                s.get('path', '') in problem_paths,
            ])
    return path


def build_recommendations(rows, prev_rows):
    """Формирует список рекомендаций из текущих и вчерашних строк."""
    recs = []
    prev_idx = {tuple(r['keys']): r for r in prev_rows}

    for r in rows:
        query, page = r['keys']
        imp = r.get('impressions', 0)
        ctr = r.get('ctr', 0)
        pos = r.get('position', 0)
        clicks = r.get('clicks', 0)

        if imp >= LOW_CTR_MIN_IMPRESSIONS and ctr < LOW_CTR_THRESHOLD:
            recs.append({
                'type': 'low_ctr', 'query': query, 'page': page,
                'clicks': clicks, 'impressions': imp, 'ctr': ctr, 'position': pos,
                'reason': f'CTR {ctr*100:.1f}% при {imp} показах — переписать title/description',
            })

        if PAGE_2_MIN <= pos <= PAGE_2_MAX and imp > 0:
            recs.append({
                'type': 'page_2', 'query': query, 'page': page,
                'clicks': clicks, 'impressions': imp, 'ctr': ctr, 'position': pos,
                'reason': f'Позиция {pos:.1f} — вторая страница, вытащить в топ-10',
            })

        prev = prev_idx.get((query, page))
        if prev:
            prev_pos = prev.get('position', 0)
            if prev_pos > 0 and pos - prev_pos >= POSITION_DROP_DELTA:
                recs.append({
                    'type': 'position_drop', 'query': query, 'page': page,
                    'clicks': clicks, 'impressions': imp, 'ctr': ctr, 'position': pos,
                    'reason': f'Позиция ухудшилась: {prev_pos:.1f} -> {pos:.1f}',
                })
    return recs


def write_recommendations_csv(recs, target_date):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f'recommendations_{target_date}.csv')

    def sort_key(r):
        host = get_host(r['page'])
        type_order = {'position_drop': 0, 'low_ctr': 1, 'page_2': 2}
        return (
            host_sort_key(host),
            type_order.get(r['type'], 99),
            -(r['impressions'] or 0),
        )

    recs_sorted = sorted(recs, key=sort_key)
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'type', 'subdomain', 'query', 'page', 'clicks',
            'impressions', 'ctr', 'position', 'reason',
        ])
        for r in recs_sorted:
            w.writerow([
                r['type'], get_host(r['page']), r['query'], r['page'],
                r['clicks'], r['impressions'],
                f'{r["ctr"]:.4f}', f'{r["position"]:.2f}', r['reason'],
            ])
    return path, len(recs_sorted)


def build_critical_block(sitemap_problems, grouped, prev_grouped):
    """Секция 'критические' для Telegram. None если проблем нет."""
    issues = []
    for p in sitemap_problems:
        path = html.escape(p['path'])
        if p['errors'] > 0:
            issues.append(f'🔴 Sitemap {path}: ошибок {p["errors"]}')
        elif p['warnings'] > 0:
            issues.append(f'⚠️ Sitemap {path}: предупреждений {p["warnings"]}')
        elif not p['last_downloaded']:
            issues.append(f'⚠️ Sitemap {path}: ещё не скачан Google')

    for host, rows in grouped.items():
        curr = aggregate(rows)
        prev = aggregate(prev_grouped.get(host, []))
        if (prev['clicks'] >= CRITICAL_CLICKS_MIN_BASELINE
                and curr['clicks'] < prev['clicks'] * CRITICAL_CLICKS_DROP_RATIO):
            issues.append(
                f'🔴 {html.escape(host)}: клики упали {prev["clicks"]} -> {curr["clicks"]}'
            )

    if not issues:
        return None
    return '<b>⚠️ Критические:</b>\n' + '\n'.join(issues)


def build_host_section(host, rows, prev_rows, top_n=20):
    curr = aggregate(rows)
    prev = aggregate(prev_rows)

    lines = [
        f'📍 <b>{html.escape(host)}</b>',
        f'Запросы: <b>{curr["queries"]}</b> {fmt_int_delta(curr["queries"], prev["queries"])} · '
        f'Клики: <b>{curr["clicks"]}</b> {fmt_int_delta(curr["clicks"], prev["clicks"])}',
        f'Показы: <b>{curr["impressions"]}</b> {fmt_int_delta(curr["impressions"], prev["impressions"])} · '
        f'Поз: <b>{curr["position"]:.2f}</b> {fmt_pos_delta(curr["position"], prev["position"])}',
    ]
    top = sorted(rows, key=lambda r: r.get('clicks', 0), reverse=True)[:top_n]
    if top:
        lines.append(f'<i>Топ-{len(top)} по кликам:</i>')
        for i, r in enumerate(top, 1):
            query = html.escape(r['keys'][0])
            lines.append(
                f'  {i}. {query} — {r.get("clicks", 0)} кл. / '
                f'{r.get("impressions", 0)} пок. / поз. {r.get("position", 0):.1f}'
            )
    return '\n'.join(lines)


def build_summary(target_date, grouped, prev_grouped, critical_block=None):
    prev_label = prev_date(target_date)
    lines = [
        f'<b>GOOGLE Search Console — {target_date}</b>',
        f'<i>Динамика vs {prev_label}. Для позиции: минус = улучшение</i>',
        '',
    ]
    if critical_block:
        lines.append(critical_block)
        lines.append('')
    for host, rows in grouped.items():
        lines.append(build_host_section(host, rows, prev_grouped.get(host, [])))
        lines.append('')
    return '\n'.join(lines).rstrip()


def main():
    target_date = resolve_date(sys.argv)
    print(f'=== Отчёт по {SITE_URL} за {target_date} ===')

    service = get_service()
    rows = fetch_rows(service, target_date)
    print(f'Получено строк: {len(rows)}')

    if not rows:
        print('Нет данных за эту дату (возможно, ещё не обновились в Search Console).')
        return

    prev = prev_date(target_date)
    prev_rows = fetch_rows(service, prev)
    print(f'За предыдущий день ({prev}): {len(prev_rows)} строк')

    grouped = group_by_host(rows)
    prev_grouped = group_by_host(prev_rows)
    print(f'Субдоменов: {len(grouped)}')
    for host, items in grouped.items():
        path = write_csv(items, target_date, host)
        print(f'  {host}: {len(items)} строк -> {os.path.basename(path)}')

    xlsx_path, total = update_keywords_xlsx(rows, target_date)
    print(f'XLSX: {xlsx_path} (всего уникальных пар ключ+страница: {total})')

    sitemaps = fetch_sitemap_status(service)
    sitemap_problems = find_sitemap_issues(sitemaps)
    errors_path = write_errors_csv(sitemaps, sitemap_problems, target_date)
    print(f'Errors CSV: {os.path.basename(errors_path)} '
          f'(sitemap-проблем: {len(sitemap_problems)}/{len(sitemaps)})')

    recs = build_recommendations(rows, prev_rows)
    recs_path, recs_count = write_recommendations_csv(recs, target_date)
    print(f'Recommendations CSV: {os.path.basename(recs_path)} '
          f'(рекомендаций: {recs_count})')

    critical_block = build_critical_block(sitemap_problems, grouped, prev_grouped)
    if critical_block:
        print('Найдены критические предупреждения — уйдут в Telegram')

    try:
        notifier.send_report([
            build_summary(target_date, grouped, prev_grouped, critical_block)
        ])
        print('Telegram: сводка отправлена')
    except Exception as e:
        print(f'Telegram: ошибка отправки — {e}')


if __name__ == '__main__':
    main()
